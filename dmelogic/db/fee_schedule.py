"""
Fee Schedule Database Module
============================
Imports and queries the NYS Medicaid DMEPOS Fee Schedule.

The fee schedule Excel file is imported once into a `fee_schedule` table
inside billing.db. Subsequent lookups are fast SQLite queries.

Usage:
    from dmelogic.db.fee_schedule import import_fee_schedule, lookup_fee

    # Import (only needed once, or when updating)
    import_fee_schedule(xlsx_path, folder_path=...)

    # Lookup
    info = lookup_fee("T4530", folder_path=...)
    # → {"fee": 0.37, "rental_fee": 0.0, "max_units": 250, ...}
"""

from __future__ import annotations

import sqlite3
from typing import Optional, Dict, Any

from dmelogic.db.base import get_connection


# ──────────────────────────────────────────────────────────
# Table creation
# ──────────────────────────────────────────────────────────

_CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS fee_schedule (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    hcpcs_code      TEXT    NOT NULL,
    description     TEXT    NOT NULL DEFAULT '',
    fee             REAL    DEFAULT 0.0,
    rental_fee      REAL    DEFAULT 0.0,
    br              TEXT    DEFAULT '',
    max_units       INTEGER DEFAULT 0,
    pa              TEXT    DEFAULT '',
    pharmacy_only   TEXT    DEFAULT '',
    change_flag     TEXT    DEFAULT '',
    effective_date  TEXT    DEFAULT '',
    UNIQUE(hcpcs_code)
);
"""

_CREATE_INDEX_SQL = """
CREATE INDEX IF NOT EXISTS idx_fee_schedule_hcpcs ON fee_schedule(hcpcs_code);
"""


def _ensure_table(conn: sqlite3.Connection) -> None:
    conn.execute(_CREATE_TABLE_SQL)
    conn.execute(_CREATE_INDEX_SQL)
    conn.commit()


# ──────────────────────────────────────────────────────────
# Import from Excel
# ──────────────────────────────────────────────────────────

def import_fee_schedule(
    xlsx_path: str,
    folder_path: Optional[str] = None,
    effective_date: str = "",
) -> int:
    """
    Import HCPCS fee schedule from a Medicaid XLSX file into billing.db.

    Expected format (NYS Medicaid DMEPOS):
      Row 1: Title
      Row 2: Effective date
      Row 3: Headers (CODE, DESCRIPTION, FEE, RENTAL FEE, BR, MAX UNITS, PA, PHARMACY ONLY, CHANGE)
      Row 4+: Data

    Returns the number of rows imported.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Auto-detect effective date from row 2 if not provided
    if not effective_date:
        row2_val = ws.cell(row=2, column=1).value or ""
        if "effective" in str(row2_val).lower():
            effective_date = str(row2_val).strip()

    # Collect rows (data starts at row 4)
    rows_to_insert = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        code = row[0]
        if not code or not str(code).strip():
            continue

        hcpcs = str(code).strip().upper()
        description = str(row[1] or "").strip()

        def _to_float(val):
            if val is None:
                return 0.0
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        def _to_int(val):
            if val is None:
                return 0
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return 0

        fee = _to_float(row[2]) if len(row) > 2 else 0.0
        rental_fee = _to_float(row[3]) if len(row) > 3 else 0.0
        br = str(row[4] or "").strip() if len(row) > 4 else ""
        max_units = _to_int(row[5]) if len(row) > 5 else 0
        pa = str(row[6] or "").strip() if len(row) > 6 else ""
        pharmacy_only = str(row[7] or "").strip() if len(row) > 7 else ""
        change_flag = str(row[8] or "").strip() if len(row) > 8 else ""

        rows_to_insert.append((
            hcpcs, description, fee, rental_fee, br,
            max_units, pa, pharmacy_only, change_flag, effective_date,
        ))

    wb.close()

    # Write to DB
    conn = get_connection("billing.db", folder_path=folder_path)
    _ensure_table(conn)

    # Clear old data and re-import
    conn.execute("DELETE FROM fee_schedule")
    conn.executemany(
        """INSERT OR REPLACE INTO fee_schedule
           (hcpcs_code, description, fee, rental_fee, br, max_units, pa, pharmacy_only, change_flag, effective_date)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        rows_to_insert,
    )
    conn.commit()
    conn.close()

    return len(rows_to_insert)


# ──────────────────────────────────────────────────────────
# Fee lookup
# ──────────────────────────────────────────────────────────

def lookup_fee(
    hcpcs: str,
    folder_path: Optional[str] = None,
    rental: bool = False,
) -> Optional[Dict[str, Any]]:
    """
    Look up the Medicaid fee for a given HCPCS code.

    Args:
        hcpcs: HCPCS code (e.g. "T4530", "A4495")
        folder_path: DB folder path override
        rental: If True, return rental_fee instead of purchase fee

    Returns dict with keys: fee, rental_fee, max_units, pa, description, pharmacy_only
    or None if not found.
    """
    if not hcpcs:
        return None

    hcpcs_clean = hcpcs.strip().upper()

    try:
        conn = get_connection("billing.db", folder_path=folder_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        # Ensure table exists
        _ensure_table(conn)

        # Exact match
        cur.execute(
            "SELECT * FROM fee_schedule WHERE hcpcs_code = ? LIMIT 1",
            (hcpcs_clean,)
        )
        row = cur.fetchone()

        if not row:
            # Try base code (strip dash suffix like A4495-NU)
            base = hcpcs_clean.split("-")[0] if "-" in hcpcs_clean else hcpcs_clean
            if base != hcpcs_clean:
                cur.execute(
                    "SELECT * FROM fee_schedule WHERE hcpcs_code = ? LIMIT 1",
                    (base,)
                )
                row = cur.fetchone()

        conn.close()

        if row:
            fee_val = row["rental_fee"] if rental and row["rental_fee"] else row["fee"]
            return {
                "fee": row["fee"] or 0.0,
                "rental_fee": row["rental_fee"] or 0.0,
                "max_units": row["max_units"] or 0,
                "pa": row["pa"] or "",
                "description": row["description"] or "",
                "pharmacy_only": row["pharmacy_only"] or "",
                "br": row["br"] or "",
                "effective_fee": fee_val or 0.0,
            }
    except Exception:
        pass

    return None


def is_fee_schedule_loaded(folder_path: Optional[str] = None) -> bool:
    """Check if the fee schedule table exists and has data."""
    try:
        conn = get_connection("billing.db", folder_path=folder_path)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM fee_schedule")
        count = cur.fetchone()[0]
        conn.close()
        return count > 0
    except Exception:
        return False


def get_fee_schedule_count(folder_path: Optional[str] = None) -> int:
    """Return the number of codes in the fee schedule."""
    try:
        conn = get_connection("billing.db", folder_path=folder_path)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM fee_schedule")
        count = cur.fetchone()[0]
        conn.close()
        return count
    except Exception:
        return 0
