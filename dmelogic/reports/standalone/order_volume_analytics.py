"""
Order Volume Analytics Report

Interactive dashboard showing order trends, status breakdowns,
new-vs-repeat patient analysis, and top HCPCS / insurance usage.
Adapted to actual DMELogic database schema (orders.db).
"""

from __future__ import annotations
import sqlite3
from collections import defaultdict
from datetime import datetime, date, timedelta
from typing import Optional, Dict, List
from pathlib import Path

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
    QFileDialog, QFrame, QComboBox, QGroupBox, QTabWidget,
    QWidget, QGridLayout, QSplitter
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor, QBrush

# Optional matplotlib for charts
try:
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False


class OrderVolumeAnalyticsReport(QDialog):
    """Order Volume Analytics / Dashboard"""

    def __init__(self, parent=None, *, orders_db_path: str = "",
                 folder_path: Optional[str] = None):
        super().__init__(parent)
        if orders_db_path:
            self.orders_db_path = Path(orders_db_path)
        elif folder_path:
            self.orders_db_path = Path(folder_path) / "orders.db"
        else:
            self.orders_db_path = Path("orders.db")

        self.period = "monthly"

        self.setWindowTitle("Order Volume Analytics")
        self.setModal(False)
        self.resize(1500, 850)
        self.setMinimumSize(1100, 600)

        self._setup_ui()
        self._generate_report()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # Title + period selector
        top = QHBoxLayout()
        hdr = QLabel("Order Volume Analytics")
        hf = QFont("Segoe UI", 14)
        hf.setBold(True)
        hdr.setFont(hf)
        hdr.setStyleSheet("color: #1565c0; padding: 8px;")
        top.addWidget(hdr)
        top.addStretch()
        top.addWidget(QLabel("Period:"))
        self.period_combo = QComboBox()
        self.period_combo.addItem("Weekly", "weekly")
        self.period_combo.addItem("Monthly", "monthly")
        self.period_combo.addItem("Quarterly", "quarterly")
        self.period_combo.addItem("Yearly", "yearly")
        self.period_combo.setCurrentIndex(1)
        self.period_combo.currentIndexChanged.connect(self._on_period_changed)
        top.addWidget(self.period_combo)
        layout.addLayout(top)

        # KPI cards
        kpi_row = QHBoxLayout()
        self.total_orders_card = self._card("Total Orders", "0", "#1565c0")
        self.avg_value_card = self._card("Avg Order Value", "$0", "#2e7d32")
        self.unique_patients_card = self._card("Unique Patients", "0", "#7b1fa2")
        self.top_status_card = self._card("Top Status", "—", "#e65100")
        for c in (self.total_orders_card, self.avg_value_card,
                  self.unique_patients_card, self.top_status_card):
            kpi_row.addWidget(c)
        layout.addLayout(kpi_row)

        # Tabs
        tabs = QTabWidget()
        tabs.setStyleSheet("""
            QTabWidget::pane { border: 1px solid #ddd; }
            QTabBar::tab {
                background: #e3f2fd; border: 1px solid #90caf9;
                padding: 8px 16px; margin-right: 2px; border-radius: 4px 4px 0 0;
            }
            QTabBar::tab:selected { background: white; font-weight: 600; }
        """)

        # Tab 1 – Volume Trend
        trend_tab = QWidget()
        trend_lo = QVBoxLayout(trend_tab)
        self.trend_table = self._make_table(
            ["Period", "Orders", "New Patients", "Repeat Patients",
             "Revenue", "Avg Value"])
        if HAS_MATPLOTLIB:
            splitter = QSplitter(Qt.Orientation.Horizontal)
            self.trend_fig = Figure(figsize=(6, 3), dpi=100)
            self.trend_canvas = FigureCanvas(self.trend_fig)
            splitter.addWidget(self.trend_canvas)
            splitter.addWidget(self.trend_table)
            splitter.setStretchFactor(0, 3)
            splitter.setStretchFactor(1, 2)
            trend_lo.addWidget(splitter)
        else:
            trend_lo.addWidget(self.trend_table)
        tabs.addTab(trend_tab, "Volume Trend")

        # Tab 2 – Status Breakdown
        status_tab = QWidget()
        status_lo = QVBoxLayout(status_tab)
        self.status_table = self._make_table(
            ["Status", "Count", "% of Total", "Revenue"])
        if HAS_MATPLOTLIB:
            splitter2 = QSplitter(Qt.Orientation.Horizontal)
            self.status_fig = Figure(figsize=(6, 4), dpi=100)
            self.status_canvas = FigureCanvas(self.status_fig)
            splitter2.addWidget(self.status_canvas)
            splitter2.addWidget(self.status_table)
            splitter2.setStretchFactor(0, 2)
            splitter2.setStretchFactor(1, 3)
            status_lo.addWidget(splitter2)
        else:
            status_lo.addWidget(self.status_table)
        tabs.addTab(status_tab, "Status Breakdown")

        # Tab 3 – Top HCPCS
        hcpcs_tab = QWidget()
        hcpcs_lo = QVBoxLayout(hcpcs_tab)
        self.hcpcs_table = self._make_table(
            ["HCPCS", "Description", "Qty Sold", "Revenue", "# Orders"])
        if HAS_MATPLOTLIB:
            splitter3 = QSplitter(Qt.Orientation.Horizontal)
            self.hcpcs_fig = Figure(figsize=(5, 3), dpi=100)
            self.hcpcs_canvas = FigureCanvas(self.hcpcs_fig)
            splitter3.addWidget(self.hcpcs_canvas)
            splitter3.addWidget(self.hcpcs_table)
            splitter3.setStretchFactor(0, 2)
            splitter3.setStretchFactor(1, 3)
            hcpcs_lo.addWidget(splitter3)
        else:
            hcpcs_lo.addWidget(self.hcpcs_table)
        tabs.addTab(hcpcs_tab, "Top HCPCS")

        # Tab 4 – Insurance Mix
        ins_tab = QWidget()
        ins_lo = QVBoxLayout(ins_tab)
        ins_hint = QLabel("💡 Double-click an insurance row to see its orders and fix errors")
        ins_hint.setStyleSheet("color: #666; font-style: italic; padding: 4px;")
        ins_lo.addWidget(ins_hint)
        self.ins_table = self._make_table(
            ["Insurance", "# Orders", "% of Total", "Revenue"])
        self.ins_table.doubleClicked.connect(self._on_insurance_double_click)
        ins_lo.addWidget(self.ins_table)
        tabs.addTab(ins_tab, "Insurance Mix")

        layout.addWidget(tabs, 1)

        # Buttons
        btn_row = QHBoxLayout()
        for label, color, slot in [
            ("Refresh", "#1976d2", self._generate_report),
            ("Export All CSV", "#388e3c", self._export_csv),
            ("Export Excel", "#1e7e34", self._export_excel),
        ]:
            b = QPushButton(label)
            b.setStyleSheet(self._btn_css(color))
            b.clicked.connect(slot)
            btn_row.addWidget(b)
        btn_row.addStretch()
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet(self._btn_css("#6c757d"))
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

    # Helpers
    def _card(self, title, value, color):
        card = QGroupBox()
        card.setStyleSheet(f"""
            QGroupBox {{
                background: white; border: 2px solid {color};
                border-radius: 8px; padding: 12px; min-width: 140px;
            }}
        """)
        lo = QVBoxLayout(card)
        tl = QLabel(title)
        tl.setStyleSheet("font-size: 10pt; font-weight: 600; color: #666;")
        lo.addWidget(tl)
        vl = QLabel(value)
        vl.setStyleSheet(f"font-size: 20pt; font-weight: bold; color: {color};")
        vl.setObjectName("value_label")
        lo.addWidget(vl)
        return card

    def _make_table(self, columns):
        tw = QTableWidget()
        tw.setColumnCount(len(columns))
        tw.setHorizontalHeaderLabels(columns)
        tw.setAlternatingRowColors(True)
        tw.setStyleSheet("""
            QTableWidget { gridline-color: #e0e0e0; background: white; }
            QTableWidget::item { padding: 4px; }
            QHeaderView::section {
                background: #1565c0; color: white;
                padding: 6px; font-weight: 600; border: none;
            }
        """)
        hdr = tw.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        tw.verticalHeader().setVisible(False)
        return tw

    @staticmethod
    def _btn_css(color):
        return f"""
            QPushButton {{
                background-color: {color}; color: white; border: none;
                padding: 8px 16px; border-radius: 4px;
                font-weight: 600; min-width: 100px;
            }}
            QPushButton:hover {{ background-color: #333; }}
            QPushButton:pressed {{ background-color: #000; }}
            QPushButton:disabled {{ background-color: #ccc; color: #888; }}
        """

    # ------------------------------------------------------------------
    # Slots
    # ------------------------------------------------------------------
    def _on_period_changed(self):
        self.period = self.period_combo.currentData()
        self._generate_report()

    # ------------------------------------------------------------------
    # Data
    # ------------------------------------------------------------------
    @staticmethod
    def _parse_date(val, fallback=None):
        if not val:
            return fallback
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except (ValueError, AttributeError):
                continue
        return fallback

    def _period_key(self, dt: date) -> str:
        if self.period == "weekly":
            iso = dt.isocalendar()
            return f"{iso[0]}-W{iso[1]:02d}"
        elif self.period == "quarterly":
            q = (dt.month - 1) // 3 + 1
            return f"{dt.year}-Q{q}"
        elif self.period == "yearly":
            return str(dt.year)
        return f"{dt.year}-{dt.month:02d}"

    def _generate_report(self):
        if not self.orders_db_path.exists():
            QMessageBox.warning(self, "Database Not Found",
                                f"Orders database not found at:\n{self.orders_db_path}")
            return
        try:
            conn = sqlite3.connect(str(self.orders_db_path))
            conn.row_factory = sqlite3.Row
            rows = conn.execute("""
                SELECT
                    o.id                                               AS order_id,
                    COALESCE(o.order_date, o.created_date, '')        AS order_date,
                    COALESCE(o.order_status, '')                       AS status,
                    COALESCE(o.primary_insurance, '')                  AS insurance,
                    COALESCE(o.patient_id, '')                         AS patient_id,
                    SUM(CAST(COALESCE(oi.total, '0') AS REAL))         AS revenue,
                    SUM(CAST(COALESCE(oi.qty, '0') AS INTEGER))        AS qty
                FROM orders o
                LEFT JOIN order_items oi ON oi.order_id = o.id
                GROUP BY o.id
            """).fetchall()
            conn.close()

            self._data_rows = [dict(r) for r in rows]
            self._compute_all()
        except Exception as exc:
            QMessageBox.critical(self, "Report Error",
                                 f"Error generating analytics:\n\n{exc}")
            import traceback; traceback.print_exc()

    # ------------------------------------------------------------------
    def _compute_all(self):
        today = date.today()
        total_orders = len(self._data_rows)
        total_revenue = sum(r["revenue"] or 0 for r in self._data_rows)
        avg_val = total_revenue / total_orders if total_orders else 0

        patients = set()
        status_counts: Dict[str, int] = defaultdict(int)
        status_revenue: Dict[str, float] = defaultdict(float)
        period_data: Dict[str, Dict] = defaultdict(
            lambda: {"orders": 0, "new": 0, "repeat": 0, "revenue": 0.0}
        )
        seen_patients: set = set()

        for r in self._data_rows:
            pid = r["patient_id"] or r["order_id"]
            patients.add(pid)
            status = (r["status"] or "Unknown").title()
            status_counts[status] += 1
            status_revenue[status] += r["revenue"] or 0

            dt = self._parse_date(r["order_date"], today)
            pk = self._period_key(dt)
            pd_ = period_data[pk]
            pd_["orders"] += 1
            pd_["revenue"] += r["revenue"] or 0
            if pid not in seen_patients:
                pd_["new"] += 1
                seen_patients.add(pid)
            else:
                pd_["repeat"] += 1

        top_status = max(status_counts, key=status_counts.get) if status_counts else "—"

        # Update KPI cards
        self._set_card(self.total_orders_card, str(total_orders))
        self._set_card(self.avg_value_card, f"${avg_val:,.0f}")
        self._set_card(self.unique_patients_card, str(len(patients)))
        self._set_card(self.top_status_card, top_status)

        # Populate tabs
        self._fill_trend(period_data)
        self._fill_status(status_counts, status_revenue, total_orders)
        self._fill_hcpcs()
        self._fill_insurance(total_orders)

    def _set_card(self, card, text):
        lbl = card.findChild(QLabel, "value_label")
        if lbl:
            lbl.setText(text)

    # ------------------------------------------------------------------
    # Trend
    # ------------------------------------------------------------------
    def _fill_trend(self, period_data):
        self.trend_table.setRowCount(0)
        sorted_periods = sorted(period_data.keys())
        for pk in sorted_periods:
            d = period_data[pk]
            row = self.trend_table.rowCount()
            self.trend_table.insertRow(row)
            avg_v = d["revenue"] / d["orders"] if d["orders"] else 0
            for col, val in enumerate([
                pk, str(d["orders"]), str(d["new"]), str(d["repeat"]),
                f"${d['revenue']:,.2f}", f"${avg_v:,.2f}"
            ]):
                self.trend_table.setItem(row, col, QTableWidgetItem(val))

        if HAS_MATPLOTLIB and sorted_periods:
            self.trend_fig.clear()
            ax = self.trend_fig.add_subplot(111)
            ords = [period_data[p]["orders"] for p in sorted_periods]
            x = range(len(sorted_periods))
            ax.bar(x, ords, color="#42a5f5", alpha=0.8)
            ax.set_xticks(list(x))
            ax.set_xticklabels(sorted_periods, rotation=45, ha="right", fontsize=7)
            ax.set_ylabel("Orders")
            ax.set_title("Order Volume by Period", fontsize=10)
            self.trend_fig.tight_layout()
            self.trend_canvas.draw()

    # ------------------------------------------------------------------
    # Status
    # ------------------------------------------------------------------
    def _fill_status(self, counts, revenue, total):
        self.status_table.setRowCount(0)
        for status in sorted(counts, key=counts.get, reverse=True):
            cnt = counts[status]
            row = self.status_table.rowCount()
            self.status_table.insertRow(row)
            pct = cnt / total * 100 if total else 0
            for col, val in enumerate([
                status, str(cnt), f"{pct:.1f}%", f"${revenue[status]:,.2f}"
            ]):
                self.status_table.setItem(row, col, QTableWidgetItem(val))

        if HAS_MATPLOTLIB and counts:
            self.status_fig.clear()
            ax = self.status_fig.add_subplot(111)
            labels = list(counts.keys())
            sizes = list(counts.values())
            colors = ["#42a5f5", "#66bb6a", "#ffa726", "#ef5350",
                      "#ab47bc", "#26c6da", "#8d6e63", "#78909c",
                      "#ec407a", "#7e57c2"]

            # Only show percentage label on slices >= 5%
            total_size = sum(sizes)

            def autopct_func(pct):
                return f"{pct:.0f}%" if pct >= 5 else ""

            wedges, texts, autotexts = ax.pie(
                sizes,
                colors=colors[:len(labels)],
                autopct=autopct_func,
                startangle=90,
                pctdistance=0.75,
                textprops={"fontsize": 8},
            )
            # Hide on-chart labels — use legend instead
            for t in texts:
                t.set_text("")
            ax.legend(
                wedges, [f"{l} ({s})" for l, s in zip(labels, sizes)],
                loc="center left", bbox_to_anchor=(1.0, 0.5),
                fontsize=8, framealpha=0.9,
            )
            ax.set_title("Status Distribution", fontsize=10, fontweight="bold")
            self.status_fig.tight_layout(rect=[0, 0, 0.75, 1])
            self.status_canvas.draw()

    # ------------------------------------------------------------------
    # HCPCS
    # ------------------------------------------------------------------
    def _fill_hcpcs(self):
        self.hcpcs_table.setRowCount(0)
        conn = sqlite3.connect(str(self.orders_db_path))
        rows = conn.execute("""
            SELECT
                COALESCE(hcpcs_code,'UNKNOWN') AS code,
                COALESCE(description, '')       AS descr,
                SUM(CAST(COALESCE(qty,'0')  AS INTEGER)) AS total_qty,
                SUM(CAST(COALESCE(total,'0') AS REAL))   AS total_rev,
                COUNT(DISTINCT order_id)                 AS num_orders
            FROM order_items
            GROUP BY hcpcs_code
            ORDER BY total_rev DESC
            LIMIT 25
        """).fetchall()
        conn.close()

        labels, vals = [], []
        for r in rows:
            code, descr, qty, rev, nords = r
            row = self.hcpcs_table.rowCount()
            self.hcpcs_table.insertRow(row)
            for col, v in enumerate([
                code, descr[:60], str(qty), f"${rev:,.2f}", str(nords)
            ]):
                self.hcpcs_table.setItem(row, col, QTableWidgetItem(v))
            labels.append(code)
            vals.append(rev)

        if HAS_MATPLOTLIB and labels:
            self.hcpcs_fig.clear()
            ax = self.hcpcs_fig.add_subplot(111)
            top = labels[:10]
            top_vals = vals[:10]
            ax.barh(range(len(top)), top_vals, color="#66bb6a", alpha=0.85)
            ax.set_yticks(range(len(top)))
            ax.set_yticklabels(top, fontsize=8)
            ax.invert_yaxis()
            ax.set_xlabel("Revenue ($)")
            ax.set_title("Top HCPCS by Revenue", fontsize=10)
            self.hcpcs_fig.tight_layout()
            self.hcpcs_canvas.draw()

    # ------------------------------------------------------------------
    # Insurance
    # ------------------------------------------------------------------
    def _fill_insurance(self, total):
        self.ins_table.setRowCount(0)
        conn = sqlite3.connect(str(self.orders_db_path))
        rows = conn.execute("""
            SELECT
                COALESCE(o.primary_insurance, 'Unknown')    AS ins,
                COUNT(*)                                     AS cnt,
                SUM(CAST(COALESCE(oi.total,'0') AS REAL))   AS rev
            FROM orders o
            LEFT JOIN order_items oi ON oi.order_id = o.id
            GROUP BY o.primary_insurance
            ORDER BY cnt DESC
        """).fetchall()
        conn.close()
        for r in rows:
            ins, cnt, rev = r
            row = self.ins_table.rowCount()
            self.ins_table.insertRow(row)
            pct = cnt / total * 100 if total else 0
            for col, v in enumerate([
                ins, str(cnt), f"{pct:.1f}%", f"${rev or 0:,.2f}"
            ]):
                self.ins_table.setItem(row, col, QTableWidgetItem(v))

    def _on_insurance_double_click(self, index):
        """Open drill-down showing all orders for the selected insurance."""
        row = index.row()
        ins_item = self.ins_table.item(row, 0)
        if not ins_item:
            return
        insurance_name = ins_item.text()
        dlg = InsuranceDrillDown(
            self, insurance_name=insurance_name,
            orders_db_path=str(self.orders_db_path)
        )
        if dlg.exec() == QDialog.DialogCode.Accepted:
            # Refresh insurance tab after edits
            total = 0
            try:
                conn = sqlite3.connect(str(self.orders_db_path))
                total = conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
                conn.close()
            except Exception:
                pass
            self._fill_insurance(total or 287)

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------
    def _export_csv(self):
        fn, _ = QFileDialog.getSaveFileName(
            self, "Export Analytics to CSV",
            f"Order_Analytics_{datetime.now():%Y%m%d_%H%M%S}.csv",
            "CSV Files (*.csv)")
        if not fn:
            return
        try:
            import csv
            with open(fn, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                # Volume Trend
                w.writerow(["--- Volume Trend ---"])
                w.writerow(["Period", "Orders", "New Patients",
                            "Repeat Patients", "Revenue", "Avg Value"])
                for i in range(self.trend_table.rowCount()):
                    w.writerow([
                        self.trend_table.item(i, c).text()
                        if self.trend_table.item(i, c) else ""
                        for c in range(self.trend_table.columnCount())
                    ])
                w.writerow([])
                # Status
                w.writerow(["--- Status Breakdown ---"])
                w.writerow(["Status", "Count", "% of Total", "Revenue"])
                for i in range(self.status_table.rowCount()):
                    w.writerow([
                        self.status_table.item(i, c).text()
                        if self.status_table.item(i, c) else ""
                        for c in range(self.status_table.columnCount())
                    ])
                w.writerow([])
                # HCPCS
                w.writerow(["--- Top HCPCS ---"])
                w.writerow(["HCPCS", "Description", "Qty", "Revenue", "# Orders"])
                for i in range(self.hcpcs_table.rowCount()):
                    w.writerow([
                        self.hcpcs_table.item(i, c).text()
                        if self.hcpcs_table.item(i, c) else ""
                        for c in range(self.hcpcs_table.columnCount())
                    ])
            QMessageBox.information(self, "Export Successful",
                                    f"Exported to:\n{fn}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", str(exc))

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(self, "Missing Module",
                                "openpyxl is not installed.\npip install openpyxl")
            return
        fn, _ = QFileDialog.getSaveFileName(
            self, "Export Analytics to Excel",
            f"Order_Analytics_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            "Excel Files (*.xlsx)")
        if not fn:
            return
        try:
            wb = openpyxl.Workbook()
            hfill = PatternFill(start_color="1565C0", end_color="1565C0",
                                fill_type="solid")
            hfont = Font(bold=True, color="FFFFFF")

            def write_sheet(ws, table):
                cols = table.columnCount()
                headers = [
                    table.horizontalHeaderItem(c).text()
                    if table.horizontalHeaderItem(c) else ""
                    for c in range(cols)
                ]
                ws.append(headers)
                for cell in ws[1]:
                    cell.fill = hfill
                    cell.font = hfont
                for i in range(table.rowCount()):
                    ws.append([
                        table.item(i, c).text()
                        if table.item(i, c) else ""
                        for c in range(cols)
                    ])
                for col in ws.columns:
                    letter = col[0].column_letter
                    mx = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[letter].width = min(mx + 2, 50)

            ws1 = wb.active
            ws1.title = "Volume Trend"
            write_sheet(ws1, self.trend_table)
            ws2 = wb.create_sheet("Status Breakdown")
            write_sheet(ws2, self.status_table)
            ws3 = wb.create_sheet("Top HCPCS")
            write_sheet(ws3, self.hcpcs_table)
            ws4 = wb.create_sheet("Insurance Mix")
            write_sheet(ws4, self.ins_table)

            wb.save(fn)
            QMessageBox.information(self, "Export Successful",
                                    f"Exported to:\n{fn}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", str(exc))


# ======================================================================
# Insurance Drill-Down Dialog
# ======================================================================
class InsuranceDrillDown(QDialog):
    """Shows all orders for a given insurance with ability to reassign."""

    def __init__(self, parent=None, *, insurance_name: str = "",
                 orders_db_path: str = ""):
        super().__init__(parent)
        self.insurance_name = insurance_name
        self.orders_db_path = orders_db_path
        self._changes_made = False

        self.setWindowTitle(f"Orders — {insurance_name}")
        self.setModal(True)
        self.resize(1200, 650)
        self.setMinimumSize(900, 400)

        self._setup_ui()
        self._load_orders()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        # Header
        hdr = QLabel(f"Orders with insurance: {self.insurance_name}")
        hf = QFont("Segoe UI", 13)
        hf.setBold(True)
        hdr.setFont(hf)
        hdr.setStyleSheet("color: #1565c0;")
        layout.addWidget(hdr)

        # Bulk-change row
        fix_row = QHBoxLayout()
        fix_row.addWidget(QLabel("Change selected orders' insurance to:"))
        self.new_insurance_input = QComboBox()
        self.new_insurance_input.setEditable(True)
        self.new_insurance_input.setMinimumWidth(250)
        self.new_insurance_input.setStyleSheet("padding: 4px; font-size: 10pt;")
        fix_row.addWidget(self.new_insurance_input)

        apply_btn = QPushButton("Apply to Selected")
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #e65100; color: white; border: none;
                padding: 6px 16px; border-radius: 4px; font-weight: 600;
            }
            QPushButton:hover { background-color: #bf360c; }
        """)
        apply_btn.clicked.connect(self._apply_change_selected)
        fix_row.addWidget(apply_btn)

        apply_all_btn = QPushButton("Apply to ALL Listed")
        apply_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #c62828; color: white; border: none;
                padding: 6px 16px; border-radius: 4px; font-weight: 600;
            }
            QPushButton:hover { background-color: #b71c1c; }
        """)
        apply_all_btn.clicked.connect(self._apply_change_all)
        fix_row.addWidget(apply_all_btn)

        fix_row.addStretch()
        layout.addLayout(fix_row)

        # Orders table
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.MultiSelection)
        cols = ["Order ID", "Order Date", "Patient", "DOB", "Phone",
                "Insurance", "Insurance ID", "Status", "Items", "Total"]
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        self.table.setStyleSheet("""
            QTableWidget { gridline-color: #e0e0e0; background: white;
                           alternate-background-color: #f8f9fa; }
            QTableWidget::item { padding: 4px; }
            QHeaderView::section {
                background: #1565c0; color: white;
                padding: 6px; font-weight: 600; border: none;
            }
        """)
        self.table.verticalHeader().setVisible(False)
        hdr_view = self.table.horizontalHeader()
        hdr_view.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        hdr_view.setStretchLastSection(True)
        layout.addWidget(self.table, 1)

        # Status bar
        self.status_label = QLabel("")
        self.status_label.setStyleSheet(
            "background: #e8f4f8; padding: 6px; border-radius: 3px; font-size: 10pt;")
        layout.addWidget(self.status_label)

        # Buttons
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #1976d2; color: white; border: none;
                padding: 8px 24px; border-radius: 4px; font-weight: 600;
            }
            QPushButton:hover { background-color: #1565c0; }
        """)
        close_btn.clicked.connect(self._on_close)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

    def _load_orders(self):
        """Load orders matching the insurance name."""
        self.table.setRowCount(0)
        try:
            conn = sqlite3.connect(self.orders_db_path)
            conn.row_factory = sqlite3.Row

            # Match: exact, empty/null for "Unknown"
            if self.insurance_name in ("Unknown", "(None)", ""):
                where = "(o.primary_insurance IS NULL OR o.primary_insurance = '')"
                params = ()
            else:
                where = "o.primary_insurance = ?"
                params = (self.insurance_name,)

            sql = f"""
                SELECT
                    o.id,
                    COALESCE(o.order_date, o.created_date, '')      AS order_date,
                    COALESCE(o.patient_last_name,'') || ', '
                        || COALESCE(o.patient_first_name,'')         AS patient_name,
                    COALESCE(o.patient_dob, '')                      AS dob,
                    COALESCE(o.patient_phone, '')                     AS phone,
                    COALESCE(o.primary_insurance,'')                  AS insurance,
                    COALESCE(o.primary_insurance_id,'')               AS insurance_id,
                    COALESCE(o.order_status,'')                       AS status,
                    GROUP_CONCAT(
                        COALESCE(oi.hcpcs_code,'') || ' ' ||
                        COALESCE(oi.description,''), ' | ')          AS items,
                    SUM(CAST(COALESCE(oi.total,'0') AS REAL))        AS total_value
                FROM orders o
                LEFT JOIN order_items oi ON oi.order_id = o.id
                WHERE {where}
                GROUP BY o.id
                ORDER BY o.order_date DESC
            """
            rows = conn.execute(sql, params).fetchall()

            # Also grab list of all distinct insurances for the combo
            all_ins = conn.execute(
                "SELECT DISTINCT primary_insurance FROM orders "
                "WHERE primary_insurance IS NOT NULL AND primary_insurance != '' "
                "ORDER BY primary_insurance"
            ).fetchall()
            conn.close()

            # Populate the combo with known insurance names
            self.new_insurance_input.clear()
            seen = set()
            for r in all_ins:
                name = r[0].strip()
                upper = name.upper()
                if upper not in seen and name:
                    self.new_insurance_input.addItem(name)
                    seen.add(upper)

            # Populate table
            for r in rows:
                row_idx = self.table.rowCount()
                self.table.insertRow(row_idx)
                items_text = (r["items"] or "").strip()
                if len(items_text) > 60:
                    items_text = items_text[:57] + "..."
                vals = [
                    f"ORD-{r['id']}",
                    r["order_date"],
                    r["patient_name"],
                    r["dob"],
                    r["phone"],
                    r["insurance"],
                    r["insurance_id"],
                    r["status"],
                    items_text,
                    f"${r['total_value'] or 0:,.2f}",
                ]
                for col, v in enumerate(vals):
                    item = QTableWidgetItem(str(v))
                    if col == 9:  # Total - right align
                        item.setTextAlignment(
                            Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.table.setItem(row_idx, col, item)

            self.status_label.setText(
                f"{len(rows)} orders with insurance \"{self.insurance_name}\"")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load orders:\n{e}")
            import traceback
            traceback.print_exc()

    def _get_selected_order_ids(self) -> list:
        """Get order IDs from selected rows."""
        ids = []
        for idx in self.table.selectionModel().selectedRows():
            item = self.table.item(idx.row(), 0)
            if item:
                oid = item.text().replace("ORD-", "")
                try:
                    ids.append(int(oid))
                except ValueError:
                    pass
        return ids

    def _apply_change_selected(self):
        """Change insurance for selected orders only."""
        new_ins = self.new_insurance_input.currentText().strip()
        if not new_ins:
            QMessageBox.warning(self, "No Insurance", "Enter or select an insurance name.")
            return
        ids = self._get_selected_order_ids()
        if not ids:
            QMessageBox.warning(self, "No Selection", "Select one or more order rows first.")
            return
        self._update_insurance(ids, new_ins)

    def _apply_change_all(self):
        """Change insurance for ALL listed orders."""
        new_ins = self.new_insurance_input.currentText().strip()
        if not new_ins:
            QMessageBox.warning(self, "No Insurance", "Enter or select an insurance name.")
            return
        count = self.table.rowCount()
        if count == 0:
            return
        reply = QMessageBox.question(
            self, "Confirm Bulk Update",
            f"Change insurance for ALL {count} orders from\n"
            f"\"{self.insurance_name}\" → \"{new_ins}\"?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        ids = []
        for r in range(count):
            item = self.table.item(r, 0)
            if item:
                oid = item.text().replace("ORD-", "")
                try:
                    ids.append(int(oid))
                except ValueError:
                    pass
        self._update_insurance(ids, new_ins)

    def _update_insurance(self, order_ids: list, new_insurance: str):
        """Update primary_insurance in orders.db for the given IDs."""
        if not order_ids:
            return
        try:
            conn = sqlite3.connect(self.orders_db_path)
            placeholders = ",".join("?" for _ in order_ids)
            conn.execute(
                f"UPDATE orders SET primary_insurance = ? WHERE id IN ({placeholders})",
                [new_insurance] + order_ids
            )
            conn.commit()
            conn.close()
            self._changes_made = True

            QMessageBox.information(
                self, "Updated",
                f"Updated {len(order_ids)} order(s) to \"{new_insurance}\"."
            )
            # Reload the view
            self._load_orders()

        except Exception as e:
            QMessageBox.critical(self, "Update Failed", f"Error:\n{e}")
            import traceback
            traceback.print_exc()

    def _on_close(self):
        if self._changes_made:
            self.accept()  # Signal parent to refresh
        else:
            self.reject()
