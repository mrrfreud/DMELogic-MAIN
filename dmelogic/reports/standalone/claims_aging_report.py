"""
Claims Aging (Accounts Receivable) Report

Shows outstanding orders grouped by age buckets for cash flow management.
Queries orders.db for orders with status Billed/Shipped/Delivered/Picked Up
that are NOT yet paid (paid = 0).  Billed amount is the SUM of order_items.total.
"""

from __future__ import annotations
import sqlite3
import os
from datetime import datetime, date
from typing import Optional, Dict, List
from pathlib import Path

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
    QFileDialog, QFrame, QTextEdit, QComboBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor, QBrush


class ClaimsAgingReport(QDialog):
    """
    Accounts Receivable (AR) Aging Report.

    Shows unpaid orders (Billed / Shipped / Delivered / Picked Up) grouped by age:
    - 0-30 days (current)
    - 31-60 days (attention needed)
    - 61-90 days (urgent)
    - 90+ days (critical)
    """

    def __init__(self, parent=None, *, billing_db_path: str = "",
                 orders_db_path: str = "", folder_path: str = ""):
        super().__init__(parent)
        # Resolve database paths ------------------------------------------
        if orders_db_path:
            self.orders_db_path = orders_db_path
        elif folder_path:
            self.orders_db_path = str(Path(folder_path) / "orders.db")
        else:
            self.orders_db_path = ""

        # Keep billing_db_path for backward compat (unused now)
        self.billing_db_path = billing_db_path or ""

        self.claims_data: List[Dict] = []
        self.summary_stats: Dict = {}
        self._current_filter = "all"  # aging bucket filter

        self.setWindowTitle("Accounts Receivable Aging Report")
        self.setModal(False)
        self.resize(1200, 750)
        self.setMinimumSize(900, 500)

        self._setup_ui()
        self._generate_report()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # Header
        header = QLabel("Accounts Receivable Aging Report")
        hf = QFont("Segoe UI", 14)
        hf.setBold(True)
        header.setFont(hf)
        header.setStyleSheet("color: #1976d2; padding: 8px;")
        layout.addWidget(header)

        # Summary frame
        self.summary_frame = QFrame()
        self.summary_frame.setFrameShape(QFrame.Shape.StyledPanel)
        self.summary_frame.setStyleSheet("""
            QFrame {
                background-color: #f5f5f5;
                border: 1px solid #ddd;
                border-radius: 6px;
                padding: 12px;
            }
        """)
        sl = QVBoxLayout(self.summary_frame)
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        self.summary_text.setMaximumHeight(200)
        self.summary_text.setStyleSheet("""
            QTextEdit {
                background-color: white;
                border: none;
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 10pt;
            }
        """)
        sl.addWidget(self.summary_text)
        layout.addWidget(self.summary_frame)

        # Filter row
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("Show:"))
        self.bucket_combo = QComboBox()
        self.bucket_combo.addItems([
            "All Claims", "0-30 Days", "31-60 Days", "61-90 Days", "90+ Days"
        ])
        self.bucket_combo.currentIndexChanged.connect(self._apply_filter)
        self.bucket_combo.setMinimumWidth(140)
        filter_row.addWidget(self.bucket_combo)
        filter_row.addStretch()
        layout.addLayout(filter_row)

        # Claims table
        tl = QLabel("Outstanding Orders Detail:")
        tl.setStyleSheet("font-weight: 600; font-size: 11pt; margin-top: 4px;")
        layout.addWidget(tl)

        self.claims_table = QTableWidget()
        self.claims_table.setColumnCount(8)
        self.claims_table.setHorizontalHeaderLabels([
            "Order ID", "Order Date", "Patient", "Insurance",
            "Billed", "Balance", "Days Old", "Status"
        ])
        self.claims_table.setAlternatingRowColors(True)
        self.claims_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #e0e0e0;
                background-color: white;
            }
            QTableWidget::item { padding: 4px; }
            QHeaderView::section {
                background-color: #1976d2;
                color: white;
                padding: 6px;
                font-weight: 600;
                border: none;
            }
        """)
        hdr = self.claims_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.claims_table.verticalHeader().setVisible(False)
        layout.addWidget(self.claims_table, 1)

        # Buttons
        btn_row = QHBoxLayout()

        refresh_btn = QPushButton("Refresh")
        refresh_btn.setStyleSheet(self._btn_css("#1976d2"))
        refresh_btn.clicked.connect(self._generate_report)
        btn_row.addWidget(refresh_btn)

        csv_btn = QPushButton("Export CSV")
        csv_btn.setStyleSheet(self._btn_css("#388e3c"))
        csv_btn.clicked.connect(self._export_csv)
        btn_row.addWidget(csv_btn)

        xl_btn = QPushButton("Export Excel")
        xl_btn.setStyleSheet(self._btn_css("#1e7e34"))
        xl_btn.clicked.connect(self._export_excel)
        btn_row.addWidget(xl_btn)

        btn_row.addStretch()

        close_btn = QPushButton("Close")
        close_btn.setStyleSheet(self._btn_css("#6c757d"))
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)

        layout.addLayout(btn_row)

    @staticmethod
    def _btn_css(color: str) -> str:
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: 600;
                min-width: 100px;
            }}
            QPushButton:hover {{ background-color: #333; }}
            QPushButton:pressed {{ background-color: #000; }}
        """

    # ------------------------------------------------------------------
    # Data
    # ------------------------------------------------------------------
    def _generate_report(self):
        if not self.orders_db_path or not os.path.exists(self.orders_db_path):
            QMessageBox.warning(
                self, "Database Not Found",
                f"Orders database not found at:\n{self.orders_db_path}\n\n"
                "Please ensure your database files are in the correct location."
            )
            return
        try:
            self.claims_data = self._fetch_claims()
            self.summary_stats = self._calc_summary()
            self._populate_summary()
            self._populate_table()
        except Exception as exc:
            QMessageBox.critical(self, "Report Error",
                                 f"Failed to generate AR report:\n\n{exc}")
            import traceback
            traceback.print_exc()

    def _fetch_claims(self) -> List[Dict]:
        """Fetch unpaid orders with status Billed/Shipped/Delivered/Picked Up.

        The billed amount is the SUM of order_items.total for each order.
        """
        conn = sqlite3.connect(self.orders_db_path)
        conn.row_factory = sqlite3.Row

        sql = """
            SELECT
                o.id                                         AS order_id,
                COALESCE(o.order_date, o.created_date)       AS order_date,
                COALESCE(o.patient_last_name, '') || ', '
                    || COALESCE(o.patient_first_name, '')     AS patient_name,
                COALESCE(o.primary_insurance, '(Unknown)')   AS insurance_name,
                COALESCE(item_totals.billed, 0)              AS billed_amount,
                o.order_status                               AS status
            FROM orders o
            LEFT JOIN (
                SELECT order_id,
                       SUM(CAST(COALESCE(total, '0') AS REAL)) AS billed
                FROM order_items
                GROUP BY order_id
            ) item_totals ON item_totals.order_id = o.id
            WHERE LOWER(COALESCE(o.order_status, ''))
                  IN ('billed', 'shipped', 'delivered', 'picked up')
              AND COALESCE(o.paid, 0) = 0
            ORDER BY o.order_date ASC
        """

        rows = conn.execute(sql).fetchall()
        conn.close()

        claims: List[Dict] = []
        today = date.today()
        for row in rows:
            cd = self._parse_date(row["order_date"], today)
            billed = float(row["billed_amount"] or 0)
            claims.append({
                "claim_id": f"ORD-{row['order_id']}",
                "claim_date": cd,
                "patient_name": row["patient_name"] or "",
                "insurance_name": row["insurance_name"],
                "billed_amount": billed,
                "paid_amount": 0.0,
                "balance": billed,
                "days_old": (today - cd).days,
                "status": row["status"],
                "denial_reason": "",
            })
        return claims

    @staticmethod
    def _parse_date(val, fallback):
        if not val:
            return fallback
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
        return fallback

    # ------------------------------------------------------------------
    # Aging summary
    # ------------------------------------------------------------------
    def _calc_summary(self) -> Dict:
        buckets = {
            "0-30":  {"count": 0, "amount": 0.0, "claims": []},
            "31-60": {"count": 0, "amount": 0.0, "claims": []},
            "61-90": {"count": 0, "amount": 0.0, "claims": []},
            "90+":   {"count": 0, "amount": 0.0, "claims": []},
        }
        total_ar = 0.0
        for c in self.claims_data:
            d = c["days_old"]
            b = c["balance"]
            total_ar += b
            key = "0-30" if d <= 30 else "31-60" if d <= 60 else "61-90" if d <= 90 else "90+"
            buckets[key]["count"] += 1
            buckets[key]["amount"] += b
            buckets[key]["claims"].append(c)
        return {"buckets": buckets, "total_ar": total_ar,
                "total_claims": len(self.claims_data)}

    def _populate_summary(self):
        s = self.summary_stats
        bk = s["buckets"]
        total = s["total_ar"]
        n = s["total_claims"]

        lines = [
            "=" * 70,
            "ACCOUNTS RECEIVABLE AGING SUMMARY",
            "=" * 70,
            "",
            f"Report Date: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}",
            f"Total Outstanding: ${total:,.2f} ({n} orders)",
            "",
            "AGING BUCKETS:",
            "-" * 70,
        ]

        labels = [
            ("0-30",  "0-30 days: ", "CURRENT"),
            ("31-60", "31-60 days:", "ATTENTION"),
            ("61-90", "61-90 days:", "URGENT"),
            ("90+",   "90+ days:  ", "CRITICAL"),
        ]
        for key, lbl, tag in labels:
            b = bk[key]
            pct = (b["amount"] / total * 100) if total else 0
            lines.append(
                f"  {lbl}  ${b['amount']:>12,.2f}  ({b['count']:>3} orders)  "
                f"{pct:>5.1f}%  {tag}"
            )
        lines.append("-" * 70)

        critical = bk["90+"]["claims"]
        if critical:
            lines += ["", "ORDERS REQUIRING IMMEDIATE FOLLOW-UP (90+ Days):", ""]
            for c in sorted(critical, key=lambda x: x["days_old"], reverse=True)[:5]:
                lines.append(
                    f"   - {c['claim_id']} - {c['patient_name']} - ${c['balance']:,.2f} - "
                    f"{c['insurance_name']} - {c['days_old']} days"
                )

        self.summary_text.setPlainText("\n".join(lines))

    # ------------------------------------------------------------------
    # Table
    # ------------------------------------------------------------------
    def _bucket_for(self, days: int) -> str:
        if days <= 30:
            return "0-30"
        if days <= 60:
            return "31-60"
        if days <= 90:
            return "61-90"
        return "90+"

    def _apply_filter(self, idx: int):
        mapping = {0: "all", 1: "0-30", 2: "31-60", 3: "61-90", 4: "90+"}
        self._current_filter = mapping.get(idx, "all")
        self._populate_table()

    def _populate_table(self):
        self.claims_table.setRowCount(0)
        filtered = self.claims_data
        if self._current_filter != "all":
            filtered = [c for c in filtered
                        if self._bucket_for(c["days_old"]) == self._current_filter]

        for claim in sorted(filtered, key=lambda x: x["days_old"], reverse=True):
            row = self.claims_table.rowCount()
            self.claims_table.insertRow(row)

            items = [
                (str(claim["claim_id"]), None),
                (claim["claim_date"].strftime("%m/%d/%Y"), None),
                (claim["patient_name"], None),
                (claim["insurance_name"], None),
                (f"${claim['billed_amount']:,.2f}",
                 Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter),
                (f"${claim['balance']:,.2f}",
                 Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter),
                (str(claim["days_old"]),
                 Qt.AlignmentFlag.AlignCenter),
                (claim["status"], None),
            ]

            days = claim["days_old"]
            if days > 90:
                bg = QColor(255, 230, 230)
            elif days > 60:
                bg = QColor(255, 243, 224)
            elif days > 30:
                bg = QColor(255, 252, 230)
            else:
                bg = QColor(230, 255, 230)

            for col, (text, align) in enumerate(items):
                item = QTableWidgetItem(text)
                if align:
                    item.setTextAlignment(align)
                item.setBackground(QBrush(bg))
                self.claims_table.setItem(row, col, item)

    # ------------------------------------------------------------------
    # Exports
    # ------------------------------------------------------------------
    def _export_csv(self):
        if not self.claims_data:
            QMessageBox.information(self, "No Data", "No claims data to export.")
            return
        fn, _ = QFileDialog.getSaveFileName(
            self, "Export AR Report to CSV",
            f"AR_Report_{datetime.now():%Y%m%d_%H%M%S}.csv",
            "CSV Files (*.csv)")
        if not fn:
            return
        try:
            import csv
            with open(fn, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["Order ID", "Order Date", "Patient", "Insurance",
                            "Billed", "Balance", "Days Old", "Status"])
                for c in sorted(self.claims_data,
                                key=lambda x: x["days_old"], reverse=True):
                    w.writerow([
                        c["claim_id"],
                        c["claim_date"].strftime("%m/%d/%Y"),
                        c["patient_name"], c["insurance_name"],
                        f"${c['billed_amount']:.2f}",
                        f"${c['balance']:.2f}",
                        c["days_old"], c["status"],
                    ])
            QMessageBox.information(self, "Export Successful",
                                    f"AR report exported to:\n{fn}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed",
                                 f"Failed to export CSV:\n{exc}")

    def _export_excel(self):
        if not self.claims_data:
            QMessageBox.information(self, "No Data", "No claims data to export.")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(
                self, "Module Not Found",
                "openpyxl is not installed.\n\nInstall with:\n"
                "pip install openpyxl")
            return
        fn, _ = QFileDialog.getSaveFileName(
            self, "Export AR Report to Excel",
            f"AR_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            "Excel Files (*.xlsx)")
        if not fn:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "AR Aging"
            headers = ["Order ID", "Order Date", "Patient", "Insurance",
                       "Billed", "Balance", "Days Old", "Status"]
            ws.append(headers)
            hfill = PatternFill(start_color="1976D2", end_color="1976D2",
                                fill_type="solid")
            hfont = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = hfill
                cell.font = hfont
                cell.alignment = Alignment(horizontal="center")

            for c in sorted(self.claims_data,
                            key=lambda x: x["days_old"], reverse=True):
                ws.append([
                    c["claim_id"],
                    c["claim_date"].strftime("%m/%d/%Y"),
                    c["patient_name"], c["insurance_name"],
                    c["billed_amount"],
                    c["balance"], c["days_old"], c["status"],
                ])
                days = c["days_old"]
                if days > 90:
                    fill = PatternFill(start_color="FFE6E6",
                                      end_color="FFE6E6", fill_type="solid")
                elif days > 60:
                    fill = PatternFill(start_color="FFF3E0",
                                      end_color="FFF3E0", fill_type="solid")
                elif days > 30:
                    fill = PatternFill(start_color="FFFCE6",
                                      end_color="FFFCE6", fill_type="solid")
                else:
                    fill = PatternFill(start_color="E6FFE6",
                                      end_color="E6FFE6", fill_type="solid")
                for cell in ws[ws.max_row]:
                    cell.fill = fill

            for col in ws.columns:
                letter = col[0].column_letter
                mx = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[letter].width = min(mx + 2, 50)

            wb.save(fn)
            QMessageBox.information(self, "Export Successful",
                                    f"AR report exported to:\n{fn}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed",
                                 f"Failed to export Excel:\n{exc}")
