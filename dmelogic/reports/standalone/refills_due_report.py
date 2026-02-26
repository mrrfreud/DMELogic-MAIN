"""
Refills Due Report - Revamped Version

Shows upcoming refills with actionable scheduling and auto-generation features.
Queries orders.db for orders with refill_due_date or day_supply in order_items.
Adapted to actual DMELogic database schema.
"""

from __future__ import annotations
import sqlite3
from datetime import datetime, date, timedelta
from typing import Optional, Dict, List, Tuple
from pathlib import Path

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
    QFileDialog, QFrame, QComboBox, QCheckBox, QGroupBox, QWidget
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor, QBrush


class RefillsDueReport(QDialog):
    """
    Refills Due Report - Shows upcoming and overdue refills.

    Pulls from orders.db using:
    - refill_due_date column on orders
    - day_supply from order_items to calculate due dates
    - Filters for Delivered/Shipped/Picked Up/Billed orders
    """

    def __init__(self, parent=None, *, orders_db_path: str = "",
                 folder_path: Optional[str] = None):
        super().__init__(parent)
        if orders_db_path:
            self.orders_db_path = Path(orders_db_path)
        elif folder_path:
            self.orders_db_path = Path(folder_path) / "orders.db"
        else:
            self.orders_db_path = Path("orders.db")

        self.refills_data: List[Dict] = []
        self.summary_stats: Dict = {}
        self.current_filter = "30_days"

        self.setWindowTitle("Refills Due Report")
        self.setModal(False)
        self.resize(1400, 750)
        self.setMinimumSize(1000, 500)

        self._setup_ui()
        self._generate_report()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # Header
        header = QLabel("Refills Due Report")
        hf = QFont("Segoe UI", 14)
        hf.setBold(True)
        header.setFont(hf)
        header.setStyleSheet("color: #9c27b0; padding: 8px;")
        layout.addWidget(header)

        # Filter row
        filter_frame = QFrame()
        filter_frame.setStyleSheet("""
            QFrame {
                background-color: #f5f5f5;
                border: 1px solid #ddd;
                border-radius: 6px;
                padding: 8px;
            }
        """)
        fl = QHBoxLayout(filter_frame)
        fl.addWidget(QLabel("View:"))

        self.filter_combo = QComboBox()
        self.filter_combo.addItem("Next 7 Days", "7_days")
        self.filter_combo.addItem("Next 30 Days", "30_days")
        self.filter_combo.addItem("Next 90 Days", "90_days")
        self.filter_combo.addItem("Overdue Only", "overdue")
        self.filter_combo.addItem("All Upcoming", "all")
        self.filter_combo.setCurrentIndex(1)
        self.filter_combo.currentIndexChanged.connect(self._on_filter_changed)
        fl.addWidget(self.filter_combo)

        fl.addSpacing(20)
        self.show_completed_check = QCheckBox("Show completed refills")
        self.show_completed_check.setChecked(False)
        self.show_completed_check.toggled.connect(self._generate_report)
        fl.addWidget(self.show_completed_check)
        fl.addStretch()
        layout.addWidget(filter_frame)

        # Summary cards
        summary_row = QHBoxLayout()
        self.overdue_card = self._card("Overdue", "0", "#dc3545")
        self.week_card = self._card("This Week", "0", "#ff9800")
        self.month_card = self._card("This Month", "0", "#2196f3")
        self.value_card = self._card("Total Value", "$0", "#4caf50")
        for c in (self.overdue_card, self.week_card, self.month_card, self.value_card):
            summary_row.addWidget(c)
        layout.addLayout(summary_row)

        # Table
        tl = QLabel("Refill Schedule:")
        tl.setStyleSheet("font-weight: 600; font-size: 11pt; margin-top: 4px;")
        layout.addWidget(tl)

        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "Order #", "Patient", "Phone", "Item / HCPCS",
            "Last Order", "Refill Due", "Days Until",
            "Qty", "Est. Value", "Status"
        ])
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget { gridline-color: #e0e0e0; background-color: white; }
            QTableWidget::item { padding: 4px; }
            QHeaderView::section {
                background-color: #9c27b0; color: white;
                padding: 6px; font-weight: 600; border: none;
            }
        """)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        layout.addWidget(self.table, 1)

        # Buttons
        btn_row = QHBoxLayout()
        for label, color, slot in [
            ("Refresh", "#1976d2", self._generate_report),
            ("Export CSV", "#388e3c", self._export_csv),
            ("Export Excel", "#1e7e34", self._export_excel),
        ]:
            b = QPushButton(label)
            b.setStyleSheet(self._btn_css(color))
            b.clicked.connect(slot)
            btn_row.addWidget(b)
        btn_row.addStretch()
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet(self._btn_css("#6c757d"))
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

    def _card(self, title: str, value: str, color: str) -> QGroupBox:
        card = QGroupBox()
        card.setStyleSheet(f"""
            QGroupBox {{
                background-color: white;
                border: 2px solid {color};
                border-radius: 8px;
                padding: 12px;
                min-width: 150px;
            }}
        """)
        lo = QVBoxLayout(card)
        tl = QLabel(title)
        tl.setStyleSheet("font-size: 10pt; font-weight: 600; color: #666;")
        lo.addWidget(tl)
        vl = QLabel(value)
        vl.setStyleSheet(f"font-size: 20pt; font-weight: bold; color: {color};")
        vl.setObjectName("value_label")
        lo.addWidget(vl)
        return card

    @staticmethod
    def _btn_css(color: str) -> str:
        return f"""
            QPushButton {{
                background-color: {color}; color: white; border: none;
                padding: 8px 16px; border-radius: 4px;
                font-weight: 600; min-width: 100px;
            }}
            QPushButton:hover {{ background-color: #333; }}
            QPushButton:pressed {{ background-color: #000; }}
            QPushButton:disabled {{ background-color: #ccc; color: #888; }}
        """

    # ------------------------------------------------------------------
    # Slots
    # ------------------------------------------------------------------
    def _on_filter_changed(self):
        self.current_filter = self.filter_combo.currentData()
        self._generate_report()

    # ------------------------------------------------------------------
    # Data
    # ------------------------------------------------------------------
    def _generate_report(self):
        if not self.orders_db_path.exists():
            QMessageBox.warning(self, "Database Not Found",
                                f"Orders database not found at:\n{self.orders_db_path}")
            return
        try:
            self.refills_data = self._fetch_refills()
            self.summary_stats = self._calc_summary()
            self._update_cards()
            self._populate_table()
        except Exception as exc:
            QMessageBox.critical(self, "Report Error",
                                 f"Failed to generate refills report:\n\n{exc}")
            import traceback; traceback.print_exc()

    @staticmethod
    def _parse_date(val, fallback):
        if not val:
            return fallback
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except (ValueError, AttributeError):
                continue
        return fallback

    def _fetch_refills(self) -> List[Dict]:
        """Fetch upcoming/overdue refills from orders.db.

        Strategy:
        1. If refill_due_date is set on the order, use it directly.
        2. Otherwise, compute due date from delivery_date + MAX(day_supply)
           from order_items.
        Only consider Delivered / Shipped / Picked Up / Billed orders
        that are NOT refill_completed.
        """
        conn = sqlite3.connect(str(self.orders_db_path))
        conn.row_factory = sqlite3.Row

        sql = """
            SELECT
                o.id                                                          AS order_id,
                COALESCE(o.patient_last_name,'') || ', '
                    || COALESCE(o.patient_first_name,'')                       AS patient_name,
                COALESCE(o.patient_phone, '')                                 AS phone,
                COALESCE(o.refill_due_date, '')                               AS refill_due_date,
                COALESCE(o.delivery_date, o.order_date, o.created_date)       AS last_date,
                COALESCE(o.order_status, '')                                  AS status,
                COALESCE(o.refill_completed, 0)                               AS refill_completed,
                COALESCE(o.refill_number, 0)                                  AS refill_number,
                MAX(CAST(COALESCE(oi.day_supply, '30') AS INTEGER))           AS day_supply,
                GROUP_CONCAT(COALESCE(oi.hcpcs_code,'') || ' ' ||
                             COALESCE(oi.description,''), ' | ')              AS items_text,
                SUM(CAST(COALESCE(oi.qty, '0') AS INTEGER))                   AS total_qty,
                SUM(CAST(COALESCE(oi.total, '0') AS REAL))                    AS total_value
            FROM orders o
            LEFT JOIN order_items oi ON oi.order_id = o.id
            WHERE LOWER(COALESCE(o.order_status,''))
                  IN ('delivered','shipped','picked up','billed')
            GROUP BY o.id
            ORDER BY o.order_date ASC
        """

        rows = conn.execute(sql).fetchall()
        conn.close()

        today = date.today()
        refills: List[Dict] = []

        for row in rows:
            # Skip completed unless checkbox checked
            if row["refill_completed"] and not self.show_completed_check.isChecked():
                continue

            last_dt = self._parse_date(row["last_date"], today)
            day_supply = max(row["day_supply"] or 30, 1)

            # Determine refill due date
            rdd_str = row["refill_due_date"]
            if rdd_str:
                due = self._parse_date(rdd_str, None)
                if due is None:
                    due = last_dt + timedelta(days=day_supply)
            else:
                due = last_dt + timedelta(days=day_supply)

            days_until = (due - today).days

            # Apply filter
            filt = self.current_filter
            if filt == "7_days" and days_until > 7:
                continue
            elif filt == "30_days" and days_until > 30:
                continue
            elif filt == "90_days" and days_until > 90:
                continue
            elif filt == "overdue" and days_until >= 0:
                continue

            value = float(row["total_value"] or 0)
            qty = int(row["total_qty"] or 0)
            items = (row["items_text"] or "").strip()
            if len(items) > 80:
                items = items[:77] + "..."

            refills.append({
                "order_id": f"ORD-{row['order_id']}",
                "patient_name": row["patient_name"],
                "phone": row["phone"],
                "items": items,
                "last_order_date": last_dt,
                "refill_due_date": due,
                "days_until": days_until,
                "quantity": qty,
                "estimated_value": value,
                "status": row["status"],
            })

        return refills

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    def _calc_summary(self) -> Dict:
        overdue = [r for r in self.refills_data if r["days_until"] < 0]
        week = [r for r in self.refills_data if 0 <= r["days_until"] <= 7]
        month = [r for r in self.refills_data if 0 <= r["days_until"] <= 30]
        total_val = sum(r["estimated_value"] for r in self.refills_data)
        return {
            "overdue": len(overdue),
            "this_week": len(week),
            "this_month": len(month),
            "total_value": total_val,
        }

    def _update_cards(self):
        s = self.summary_stats
        for card, key, fmt in [
            (self.overdue_card, "overdue", str),
            (self.week_card, "this_week", str),
            (self.month_card, "this_month", str),
            (self.value_card, "total_value", lambda v: f"${v:,.0f}"),
        ]:
            lbl = card.findChild(QLabel, "value_label")
            if lbl:
                lbl.setText(fmt(s[key]))

    # ------------------------------------------------------------------
    # Table
    # ------------------------------------------------------------------
    def _populate_table(self):
        self.table.setRowCount(0)
        for ref in sorted(self.refills_data, key=lambda x: x["days_until"]):
            row = self.table.rowCount()
            self.table.insertRow(row)

            days = ref["days_until"]
            days_text = f"{days} (OVERDUE)" if days < 0 else str(days)
            status_text = ("OVERDUE" if days < 0
                           else "Ready" if days <= 7
                           else "Scheduled")

            vals = [
                (ref["order_id"], None),
                (ref["patient_name"], None),
                (ref["phone"], None),
                (ref["items"], None),
                (ref["last_order_date"].strftime("%m/%d/%Y"), None),
                (ref["refill_due_date"].strftime("%m/%d/%Y"), None),
                (days_text, Qt.AlignmentFlag.AlignCenter),
                (str(ref["quantity"]), Qt.AlignmentFlag.AlignCenter),
                (f"${ref['estimated_value']:,.2f}",
                 Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter),
                (status_text, None),
            ]

            if days < 0:
                bg = QColor(255, 230, 230)
            elif days <= 7:
                bg = QColor(255, 243, 224)
            elif days <= 30:
                bg = QColor(255, 252, 230)
            else:
                bg = QColor(240, 248, 255)

            for col, (text, align) in enumerate(vals):
                item = QTableWidgetItem(text)
                if align:
                    item.setTextAlignment(align)
                item.setBackground(QBrush(bg))
                self.table.setItem(row, col, item)

    # ------------------------------------------------------------------
    # Exports
    # ------------------------------------------------------------------
    def _export_csv(self):
        if not self.refills_data:
            QMessageBox.information(self, "No Data", "No refills to export.")
            return
        fn, _ = QFileDialog.getSaveFileName(
            self, "Export Refills to CSV",
            f"Refills_Report_{datetime.now():%Y%m%d_%H%M%S}.csv",
            "CSV Files (*.csv)")
        if not fn:
            return
        try:
            import csv
            with open(fn, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["Order #", "Patient", "Phone", "Items",
                            "Last Order", "Refill Due", "Days Until",
                            "Qty", "Est. Value", "Status"])
                for r in sorted(self.refills_data, key=lambda x: x["days_until"]):
                    w.writerow([
                        r["order_id"], r["patient_name"], r["phone"],
                        r["items"],
                        r["last_order_date"].strftime("%m/%d/%Y"),
                        r["refill_due_date"].strftime("%m/%d/%Y"),
                        r["days_until"], r["quantity"],
                        f"${r['estimated_value']:.2f}",
                        "OVERDUE" if r["days_until"] < 0 else "Scheduled",
                    ])
            QMessageBox.information(self, "Export Successful",
                                    f"Exported to:\n{fn}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", str(exc))

    def _export_excel(self):
        if not self.refills_data:
            QMessageBox.information(self, "No Data", "No refills to export.")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(self, "Module Not Found",
                                "openpyxl is not installed.\npip install openpyxl")
            return
        fn, _ = QFileDialog.getSaveFileName(
            self, "Export Refills to Excel",
            f"Refills_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            "Excel Files (*.xlsx)")
        if not fn:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Refills Due"
            headers = ["Order #", "Patient", "Phone", "Items",
                       "Last Order", "Refill Due", "Days Until",
                       "Qty", "Est. Value", "Status"]
            ws.append(headers)
            hfill = PatternFill(start_color="9C27B0", end_color="9C27B0",
                                fill_type="solid")
            hfont = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = hfill
                cell.font = hfont
                cell.alignment = Alignment(horizontal="center")

            for r in sorted(self.refills_data, key=lambda x: x["days_until"]):
                days = r["days_until"]
                ws.append([
                    r["order_id"], r["patient_name"], r["phone"],
                    r["items"],
                    r["last_order_date"].strftime("%m/%d/%Y"),
                    r["refill_due_date"].strftime("%m/%d/%Y"),
                    days, r["quantity"], r["estimated_value"],
                    "OVERDUE" if days < 0 else "Scheduled",
                ])
                if days < 0:
                    fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6",
                                      fill_type="solid")
                elif days <= 7:
                    fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0",
                                      fill_type="solid")
                elif days <= 30:
                    fill = PatternFill(start_color="FFFCE6", end_color="FFFCE6",
                                      fill_type="solid")
                else:
                    fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF",
                                      fill_type="solid")
                for cell in ws[ws.max_row]:
                    cell.fill = fill

            for col in ws.columns:
                letter = col[0].column_letter
                mx = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[letter].width = min(mx + 2, 50)
            wb.save(fn)
            QMessageBox.information(self, "Export Successful",
                                    f"Exported to:\n{fn}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", str(exc))
